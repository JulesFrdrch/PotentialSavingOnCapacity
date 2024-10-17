import os
import pandas as pd
from pyomo.core import ConcreteModel
from scipy.stats import norm
from pyomo.environ import *
from utils import *
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import pandas as pd
from openpyxl.workbook import Workbook
import matplotlib.pyplot as plt
# from _import_geographic_data import *
from matplotlib.offsetbox import AnchoredText
# from _import_optimization_files import *
import time
from pyomo.core.util import quicksum
from datetime import datetime
"""Python Dateien einfügen"""
import Ausgabe as ta
from functions import *
from missing_functions import *
from Ausgabe import *
from plots import *
#from _optimization_utils import write_output_files
from plots import *
#from _optimization_utils import *
from openpyxl import Workbook
import os

"""Bibliotheken einfügen"""
import pandas as pd
import numpy as np
import time
from termcolor import colored
from pyomo.util.model_size import build_model_size_report
from scipy.stats import norm
import math
import csv
import geopandas as gpd
from shapely import wkt
from shapely.geometry import MultiLineString, Point
from pyproj import Geod
import pickle
import matplotlib.pyplot as plt
from matplotlib.offsetbox import AnchoredText
import pyomo
from pyomo.core.util import quicksum
from pyomo.environ import *

"""Results Files"""

def write_output_file_Fleet_infos(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Flotten Details"""
    fleet_specifics = pd.DataFrame()
    for f in model.nb_fleet:  # model.nb_fleet = range(0, nb_fleets)
        d = {}
        d["fleet_id"] = f
        d["charge_cap"] = model.fleet_charge_cap[f]
        d["batt_cap"] = model.fleet_batt_cap[f]
        d["d_spec"] = model.fleet_d_spec[f]
        d["incoming"] = sum(model.fleet_incoming[f].values())
        if sum(model.fleet_incoming[f].values()) - np.sum(results["n_arrived_vehicles"][:, :, f]) > 0.01:
            d["all_arrived"] = False
        else:
            d["all_arrived"] = True

        d["arrival_SOC"] = np.sum(results["Q_arrived_vehicles"][:, :, f]) / (
                    model.fleet_charge_cap[f] * np.sum(results["n_arrived_vehicles"][:, :, f]))

        fleet_specifics = fleet_specifics.append(d, ignore_index=True)

    fleet_specifics.to_csv("results/" + "Fleet_Infos" + "(" + time_of_optimization + ")" ".csv")

def write_output_file_Charging(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Alle Charging Variablen"""
    charging = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_charge1 at t=" + str(t)] = np.sum(results["n_charge1"][t, c, :])
            d["n_output_charged1 t=" + str(t)] = np.sum(results["n_output_charged1"][t, c, :])
            d["n_finished_charge1 at t=" + str(t)] = np.sum(results["n_finished_charge1"][t, c, :])
            d["n_charge2 at t=" + str(t)] = np.sum(results["n_charge2"][t, c, :])
            d["n_output_charged2 t=" + str(t)] = np.sum(results["n_output_charged2"][t, c, :])
            d["n_finished_charge2 at t=" + str(t)] = np.sum(results["n_finished_charge2"][t, c, :])
            d["n_charge3 at t=" + str(t)] = np.sum(results["n_charge3"][t, c, :])
            d["n_output_charged3 t=" + str(t)] = np.sum(results["n_output_charged3"][t, c, :])
            d["n_finished_charge3 at t=" + str(t)] = np.sum(results["n_finished_charge3"][t, c, :])
            d["unused_capacity at t=" + str(t)] = np.sum(results["unused_capacity"][t, c, :])

        charging = charging.append(d, ignore_index=True)

    charging.to_excel("results/" + "Charging" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_Bewegung(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Alle Bewegungs Variablen"""
    movement = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_in at t=" + str(t)] = np.sum(results["n_in"][t, c, :])
            d["n_pass at t=" + str(t)] = np.sum(results["n_pass"][t, c, :])
            d["n_in_wait_charge at t=" + str(t)] = np.sum(results["n_in_wait_charge"][t, c, :])

        movement = movement.append(d, ignore_index=True)

    movement.to_excel("results/" + "Bewegungsdaten" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_n_in_wait_charge(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur n_in_wait_charge"""
    inCS = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_in_wait_charge at t=" + str(t)] = np.sum(results["n_in_wait_charge"][t, c, :])

        inCS = inCS.append(d, ignore_index=True)

    inCS.to_excel("results/" + "n_in_wait_charge" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_n_pass(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur n_pass"""
    passing = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_pass at t=" + str(t)] = np.sum(results["n_pass"][t, c, :])

        passing = passing.append(d, ignore_index=True)

    passing.to_excel("results/" + "n_pass" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_n_in(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur n_in"""
    entering = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_in at t=" + str(t)] = np.sum(results["n_in"][t, c, :])

        entering = entering.append(d, ignore_index=True)

    entering.to_excel("results/" + "n_in" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_incoming_vehicles(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur n_incoming_vehicles"""
    departing = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_incoming_vehicles at t=" + str(t)] = np.sum(results["n_incoming_vehicles"][t, c, :])

        departing = departing.append(d, ignore_index=True)

    departing.to_excel("results/" + "incoming_vehicles" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_arrived_vehicles(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur n_arriving_vehicles"""
    arriving = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_arrived_vehicles at t=" + str(t)] = np.sum(results["n_arrived_vehicles"][t, c, :])

        arriving = arriving.append(d, ignore_index=True)

    arriving.to_excel("results/" + "arrived_vehicles" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def write_output_file_charging_stations(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Gesamt Charging Cell Ausgabe"""  # model.nb_cell = range(0, nb_cells)
    cs_specifics = pd.DataFrame()
    inds_of_charging_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] > 0]
    for ij in range(0, len(inds_of_charging_cells)):
        d = {}
        c = inds_of_charging_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        for t in model.nb_timestep:
            d["n_in_wait_charge at t=" + str(t)] = np.sum(results["n_in_wait_charge"][t, c, :])
            d["Waiting Vehicles at t=" + str(t)] = np.sum(results["n_wait"][t, c, :]) + np.sum(
                results["n_wait_charge_next"][t, c, :])
            d["Charging Vehicles at t=" + str(t)] = np.sum(results["n_charge1"][t, c, :]) + np.sum(
                results["n_charge2"][t, c, :]) + np.sum(results["n_charge3"][t, c, :])
            d["Energy Charged at t=" + str(t)] = np.sum(results["E_charge1"][t, c, :]) + np.sum(
                results["E_charge2"][t, c, :]) + np.sum(results["E_charge3"][t, c, :])
            # d["Passenger Vehicles at t=" + str(t)] = np.sum(results["n_pass"][t, c, :]) + np.sum(results["n_pass"][t, c, :]) + np.sum(results["n_pass"][t, c, :])
            d["n_in at t=" + str(t)] = np.sum(results["n_in"][t, c, :])  # nur bei 2
            d["n_incoming_vehicles at t=" + str(t)] = np.sum(results["n_incoming_vehicles"][t, c, :])
            d["n_exit at t=" + str(t)] = np.sum(results["n_exit"][t, c, :])  # nur bei 2
            d["n_arrived_vehicles at t=" + str(t)] = np.sum(results["n_arrived_vehicles"][t, c, :])
            d["n_pass at t=" + str(t)] = np.sum(results["n_pass"][t, c, :])

        cs_specifics = cs_specifics.append(d, ignore_index=True)

    cs_specifics.to_excel("results/" + "charging_station_infos" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def get_variables_from_model2(model):                                                                                        #Mit meinen Results
    n_wait_charge_next = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_charge1 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )

    n_charge2 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )

    n_arrived_vehicles = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    E_charge1 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )

    Q_arrived_vehicles = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_wait = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    E_charge2 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )

    E_charge3 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_charge3 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_arrived_vehicles = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_pass = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_in = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_exit = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_incoming_vehicles = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_in_wait_charge = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_arrived_vehicles = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_charge1 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_output_charged1 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_finished_charge1 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_charge2 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_output_charged2 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_finished_charge2 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_charge3 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_output_charged3 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    n_finished_charge3 = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    unused_capacity = np.zeros(
        [
            len(model.nb_timestep),
            len(model.nb_cell),
            len(model.nb_fleet),
        ]
    )
    for t in model.nb_timestep:
        for c in model.nb_cell:
            for f in model.nb_fleet:
                if (t, c, f) in model.key_set:

                    Q_arrived_vehicles[t, c, f] = model.Q_arrived_vehicles[
                        t, c, f
                    ].value
                    n_pass[t, c, f] = model.n_pass[t, c, f].value
                    n_in[t, c, f] = model.n_in[t, c, f].value
                    n_exit[t, c, f] = model.n_exit[t, c, f].value
                    n_incoming_vehicles[t, c, f] = model.n_incoming_vehicles[t, c, f].value
                    n_arrived_vehicles[t, c, f] = model.n_arrived_vehicles[t, c, f].value

                    if (t, c, f) in model.charging_cells_key_set:
                        n_wait[t, c, f] = model.n_wait[t, c, f].value
                        unused_capacity[t, c] = model.unused_capacity[t, c].value
                        n_wait_charge_next[t, c, f] = model.n_wait_charge_next[t, c, f].value
                        n_charge1[t, c, f] = model.n_charge1[t, c, f].value
                        n_charge2[t, c, f] = model.n_charge2[t, c, f].value
                        n_charge3[t, c, f] = model.n_charge3[t, c, f].value
                        E_charge1[t, c, f] = model.E_charge1[t, c, f].value
                        E_charge2[t, c, f] = model.E_charge2[t, c, f].value
                        E_charge3[t, c, f] = model.E_charge3[t, c, f].value
                        n_in_wait_charge[t, c, f] = model.n_in_wait_charge[t, c, f].value
                        n_charge1[t, c, f] = model.n_charge1[t, c, f].value
                        n_output_charged1[t, c, f] = model.n_output_charged1[t, c, f].value
                        n_finished_charge1[t, c, f] = model.n_finished_charge1[t, c, f].value
                        n_charge2[t, c, f] = model.n_charge2[t, c, f].value
                        n_output_charged2[t, c, f] = model.n_output_charged2[t, c, f].value
                        n_finished_charge2[t, c, f] = model.n_finished_charge2[t, c, f].value
                        n_charge3[t, c, f] = model.n_charge3[t, c, f].value
                        n_output_charged3[t, c, f] = model.n_output_charged3[t, c, f].value
                        n_finished_charge3[t, c, f] = model.n_finished_charge3[t, c, f].value

    return {
        "n_arrived_vehicles": n_arrived_vehicles,
        "unused_capacity": unused_capacity,
        "Q_arrived_vehicles": Q_arrived_vehicles,
        "n_wait": n_wait,
        "n_wait_charge_next": n_wait_charge_next,
        "n_charge1": n_charge1,
        "n_charge2": n_charge2,
        "n_charge3": n_charge3,
        "E_charge1": E_charge1,
        "E_charge2": E_charge2,
        "E_charge3": E_charge3,
        "n_pass": n_pass,
        "n_in": n_in,
        "n_exit": n_exit,
        "n_incoming_vehicles": n_incoming_vehicles,
        "n_arrived_vehicles": n_arrived_vehicles,
        "n_pass": n_pass,
        "n_in_wait_charge": n_in_wait_charge,
        "n_charge1": n_charge1,
        "n_output_charged1": n_output_charged1,
        "n_finished_charge1": n_finished_charge1,
        "n_charge2": n_charge2,
        "n_output_charged2": n_output_charged2,
        "n_finished_charge2": n_finished_charge2,
        "n_charge3": n_charge3,
        "n_output_charged3": n_output_charged3,
        "n_finished_charge3": n_finished_charge3
    }


def export_energy_comparison_to_excel(charging_model, time_of_optimization):
    data = []

    # Überprüfung der geladenen Energie und maximal mögliche Energie pro Zeitschritt und Zelle
    for t in charging_model.nb_timestep:
        for c in charging_model.nb_cell:
            if charging_model.cell_charging_cap[c] > 0:
                total_energy_charged = 0
                for f in charging_model.nb_fleet:
                    if (t, c, f) in charging_model.E_charge1:
                        total_energy_charged += (
                            charging_model.E_charge1[t, c, f].value +
                            charging_model.E_charge2[t, c, f].value +
                            charging_model.E_charge3[t, c, f].value
                        )
                max_energy_possible = charging_model.cell_charging_cap[c] * charging_model.time_resolution
                utilization_percentage = (total_energy_charged / max_energy_possible) * 100 if max_energy_possible > 0 else 0
                data.append([t, c, total_energy_charged, max_energy_possible, charging_model.cell_charging_cap[c], utilization_percentage])

    # Erstellen eines DataFrames
    df = pd.DataFrame(data, columns=["Zeitschritt", "Zelle", "Geladene Energie (kWh)", "Maximale mögliche Energie (kWh)", "Gesamtkapazität der Zelle (kW)", "Auslastung (%)"])

    # Exportieren des DataFrames in eine Excel-Datei
    excel_filename = f"results/energy_comparison_{time_of_optimization}.xlsx"
    df.to_excel(excel_filename, index=False)


def write_output_file_Energy_charged_each(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur charged"""
    total_charged_each = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]

        for t in model.nb_timestep:
            d["Energy Charged1 at t=" + str(t)] = np.sum(results["E_charge1"][t, c, :])
            d["Energy Charged2 at t=" + str(t)] = np.sum(results["E_charge2"][t, c, :])
            d["Energy Charged3 at t=" + str(t)] = np.sum(results["E_charge3"][t, c, :])

        total_charged_each = total_charged_each.append(d, ignore_index=True)

    total_charged_each.to_excel("results/" + "Energy_charged_each" + "(" + time_of_optimization + ")" ".xlsx", index=False)

def print_cell_info(cells_df):
    """
    Funktion, die eine Auflistung aller Zellen zusammen mit den Informationen: length, has_cs, und capacity ausgibt.

    :param cells_df: DataFrame, der die Zelleninformationen enthält
    """
    print("Liste der eingelesenen Zellen und ihre Eigenschaften:")
    print("="*50)
    for index, row in cells_df.iterrows():
        cell_id = row['cell_id'] if 'cell_id' in row else index  # Annahme: Es gibt eine 'cell_id'-Spalte, andernfalls Index verwenden
        length = row.get('length', 'N/A')
        has_cs = row.get('has_cs', 'N/A')
        capacity = row.get('capacity', 'N/A')
        print(f"Zelle ID: {cell_id}, Länge: {length}, Hat Ladestation: {has_cs}, Kapazität: {capacity}")
    print("="*50)



def write_output_file_Energy_charged(model, time_of_optimization, filename):
    results = get_variables_from_model2(model)

    """Nur charged"""
    total_charged = pd.DataFrame()
    inds_of_all_cells = [ind for ind in model.nb_cell if model.cell_charging_cap[ind] >= 0]
    for ij in range(0, len(inds_of_all_cells)):
        d = {}
        c = inds_of_all_cells[ij]
        d["Cell ID"] = c
        d["Capacity"] = model.cell_charging_cap[c]
        d["Max_Capacity"] = 0

        for t in model.nb_timestep:
            d["Energy Charged at t=" + str(t)] = (
                np.sum(results["E_charge1"][t, c, :])
                + np.sum(results["E_charge2"][t, c, :])
                + np.sum(results["E_charge3"][t, c, :])
            )

        total_charged = total_charged.append(d, ignore_index=True)

    for idx in range(91):
        row_list = total_charged.loc[idx, :].values.flatten().tolist()
        row_list = row_list[3:]
        total_charged.loc[idx, 'Max_Capacity'] = max(row_list)

    stefan = total_charged.loc[:, 'Max_Capacity']
    total_charged.to_excel(f"results/Energy_charged({time_of_optimization}).xlsx", index=False)

    # Erstellen des Diagramms für die Summe der "Energy Charged at t"
    timesteps = model.nb_timestep
    total_energy_charged = []

    for t in timesteps:
        total_energy_t = 0
        for c in inds_of_all_cells:
            total_energy_t += (
                np.sum(results["E_charge1"][t, c, :])
                + np.sum(results["E_charge2"][t, c, :])
                + np.sum(results["E_charge3"][t, c, :])
            )
        total_energy_charged.append(total_energy_t)

    # Aufruf der Funktion zur Erstellung des Plots für die gesamte geladene Energie
    plot_total_energy_charged(timesteps, total_energy_charged, time_of_optimization)

    # Visualisierung der geladenen Energie für ausgewählte Zellen
    selected_cells = [1, 2, 3, 4, 7, 11, 13, 14]  # Beispielhafte Liste der ausgewählten Zellen, anpassen nach Bedarf

    # Aufruf der Funktion zur Erstellung des Plots für ausgewählte Zellen
    plot_energy_charged_for_selected_cells(timesteps, results, inds_of_all_cells, selected_cells, time_of_optimization)





def write_fleet_energy_and_vehicle_charging_details_to_xlsx(model: ConcreteModel, fleet_id: int, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die Energie- und Fahrzeuganzahl einer bestimmten Flotte (fleet_id) über alle Zeitschritte und Zellen mit Ladestationen (charging_cells_key_set) ausgibt.
    Zusätzlich wird der State of Charge (SOC) für in_wait_charge, wait und in_charge vehicles berechnet und ausgegeben.
    Alle Werte stehen in einer Zeile.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param fleet_id: Die ID der Flotte, für die die Daten ausgegeben werden sollen.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    # Sicherstellen, dass das Verzeichnis 'results' existiert
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # Erstelle den vollständigen Pfad zur Datei
    file_path = os.path.join(results_dir, filename)

    # Erstelle das Excel-Workbook und aktiviere das Arbeitsblatt
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Fleet_{fleet_id}_Charging_Details"

    # Schreibe die Header
    headers = [
        'Zeitschritt', 'Zelle', 'Flotten-id',
        'n_in_wait_charge', 'Q_in_charge_wait', 'SOC_in_wait_charge',
        'n_wait', 'Q_wait', 'SOC_wait',
        'n_in_charge', 'Q_in_charge', 'SOC_in_charge',
        'n_in_wait', 'Q_in_wait', 'SOC_in_wait',
        'n_finished_charging', 'Q_finished_charging', 'SOC_finished_charging'
    ]
    sheet.append(headers)

    # Iteriere über alle Zeitschritte und Zellen mit Ladestationen für die angegebene Flotte
    for t, c, f in model.charging_cells_key_set:
        if f == fleet_id:
            # Erfasse die Werte von Q_in_charge_wait, Q_wait, Q_in_charge, Q_in_wait
            Q_in_charge_wait = model.Q_in_charge_wait[t, c, fleet_id].value
            Q_wait = model.Q_wait[t, c, fleet_id].value
            Q_in_charge = model.Q_in_charge[t, c, fleet_id].value
            Q_in_wait = model.Q_in_wait[t, c, fleet_id].value
            Q_finished_charging = model.Q_finished_charging[t, c, fleet_id].value

            # Erfasse die entsprechenden n-Werte
            n_in_wait_charge = model.n_in_wait_charge[t, c, fleet_id].value
            n_wait = model.n_wait[t, c, fleet_id].value
            n_in_charge = model.n_in_charge[t, c, fleet_id].value
            n_in_wait = model.n_in_wait[t, c, fleet_id].value
            n_finished_charging = model.n_finished_charging[t, c, fleet_id].value

            # Berechne SOC-Werte, wenn n-Werte > 0 sind, um Division durch 0 zu vermeiden
            SOC_in_wait_charge = Q_in_charge_wait / n_in_wait_charge if n_in_wait_charge > 0 else 0
            SOC_wait = Q_wait / n_wait if n_wait > 0 else 0
            SOC_in_charge = Q_in_charge / n_in_charge if n_in_charge > 0 else 0
            SOC_in_wait = Q_in_wait / n_in_wait if n_in_wait > 0 else 0
            SOC_finished_charging = Q_finished_charging / n_finished_charging if n_finished_charging > 0 else 0

            # Schreibe alle Werte in eine Zeile
            sheet.append([
                t, c, fleet_id,
                n_in_wait_charge, Q_in_charge_wait, SOC_in_wait_charge,
                n_wait, Q_wait, SOC_wait,
                n_in_charge, Q_in_charge, SOC_in_charge,
                n_in_wait, Q_in_wait, SOC_in_wait,
                n_finished_charging, Q_finished_charging, SOC_finished_charging
            ])

    # Speichere das Workbook als XLSX-Datei
    workbook.save(file_path)

    #print(f"Die detaillierten Lade-Daten der Flotte {fleet_id} wurden in {file_path} gespeichert.")


def write_combined_fleet_0_to_9_energy_and_vehicle_charging_details_to_xlsx(model: ConcreteModel, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die kombinierten Energie- und Fahrzeuganzahlen der Flotten 0 bis 9 über alle Zeitschritte und Zellen mit Ladestationen (charging_cells_key_set) ausgibt.
    Zusätzlich wird der State of Charge (SOC) für in_wait_charge, wait, in_charge und finished_charging vehicles berechnet und ausgegeben.
    Alle Werte werden summiert und stehen in einer Zeile.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    # Sicherstellen, dass das Verzeichnis 'results' existiert
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # Erstelle den vollständigen Pfad zur Datei
    file_path = os.path.join(results_dir, filename)

    # Erstelle das Excel-Workbook und aktiviere das Arbeitsblatt
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fleet_0_to_9_Charging"

    # Schreibe die Header
    headers = [
        'Zeitschritt', 'Zelle',
        'n_in_wait_charge', 'Q_in_charge_wait', 'SOC_in_wait_charge',
        'n_wait', 'Q_wait', 'SOC_wait',
        'n_in_charge', 'Q_in_charge', 'SOC_in_charge',
        'n_in_wait', 'Q_in_wait', 'SOC_in_wait',
        'n_finished_charging', 'Q_finished_charging', 'SOC_finished_charging'
    ]
    sheet.append(headers)

    # Iteriere über alle Zeitschritte und Zellen mit Ladestationen
    for t in model.nb_timestep:
        for c in model.nb_cell:
            # Initialisiere die Summen für die Flotten 0 bis 9
            total_n_in_wait_charge = total_Q_in_charge_wait = total_n_wait = total_Q_wait = 0
            total_n_in_charge = total_Q_in_charge = total_n_in_wait = total_Q_in_wait = 0
            total_n_finished_charging = total_Q_finished_charging = 0

            # Iteriere über die Flotten 0 bis 9 und summiere die Werte
            for fleet_id in range(0, 10):
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    # Erfasse die Werte für diese Flotte
                    Q_in_charge_wait = model.Q_in_charge_wait[t, c, fleet_id].value
                    Q_wait = model.Q_wait[t, c, fleet_id].value
                    Q_in_charge = model.Q_in_charge[t, c, fleet_id].value
                    Q_in_wait = model.Q_in_wait[t, c, fleet_id].value
                    Q_finished_charging = model.Q_finished_charging[t, c, fleet_id].value

                    n_in_wait_charge = model.n_in_wait_charge[t, c, fleet_id].value
                    n_wait = model.n_wait[t, c, fleet_id].value
                    n_in_charge = model.n_in_charge[t, c, fleet_id].value
                    n_in_wait = model.n_in_wait[t, c, fleet_id].value
                    n_finished_charging = model.n_finished_charging[t, c, fleet_id].value

                    # Addiere die Werte zur Gesamtsumme
                    total_n_in_wait_charge += n_in_wait_charge
                    total_Q_in_charge_wait += Q_in_charge_wait
                    total_n_wait += n_wait
                    total_Q_wait += Q_wait
                    total_n_in_charge += n_in_charge
                    total_Q_in_charge += Q_in_charge
                    total_n_in_wait += n_in_wait
                    total_Q_in_wait += Q_in_wait
                    total_n_finished_charging += n_finished_charging
                    total_Q_finished_charging += Q_finished_charging

            # Berechne die SOC-Werte, wenn n-Werte > 0 sind, um Division durch 0 zu vermeiden
            SOC_in_wait_charge = total_Q_in_charge_wait / total_n_in_wait_charge if total_n_in_wait_charge > 0 else 0
            SOC_wait = total_Q_wait / total_n_wait if total_n_wait > 0 else 0
            SOC_in_charge = total_Q_in_charge / total_n_in_charge if total_n_in_charge > 0 else 0
            SOC_in_wait = total_Q_in_wait / total_n_in_wait if total_n_in_wait > 0 else 0
            SOC_finished_charging = total_Q_finished_charging / total_n_finished_charging if total_n_finished_charging > 0 else 0

            # Schreibe die summierten Werte in eine Zeile
            sheet.append([
                t, c,
                total_n_in_wait_charge, total_Q_in_charge_wait, SOC_in_wait_charge,
                total_n_wait, total_Q_wait, SOC_wait,
                total_n_in_charge, total_Q_in_charge, SOC_in_charge,
                total_n_in_wait, total_Q_in_wait, SOC_in_wait,
                total_n_finished_charging, total_Q_finished_charging, SOC_finished_charging
            ])

    # Speichere das Workbook als XLSX-Datei
    workbook.save(file_path)

    print(f"Die kombinierten Lade-Daten der Flotten 0 bis 9 wurden in {file_path} gespeichert.")




def write_combined_fleet_10_to_19_energy_and_vehicle_charging_details_to_xlsx(model: ConcreteModel, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die kombinierten Energie- und Fahrzeuganzahlen der Flotten 10 bis 19 über alle Zeitschritte und Zellen mit Ladestationen (charging_cells_key_set) ausgibt.
    Zusätzlich wird der State of Charge (SOC) für in_wait_charge, wait, in_charge und finished_charging vehicles berechnet und ausgegeben.
    Alle Werte werden summiert und stehen in einer Zeile.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    # Sicherstellen, dass das Verzeichnis 'results' existiert
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # Erstelle den vollständigen Pfad zur Datei
    file_path = os.path.join(results_dir, filename)

    # Erstelle das Excel-Workbook und aktiviere das Arbeitsblatt
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fleet_10_to_19_Charging"

    # Schreibe die Header
    headers = [
        'Zeitschritt', 'Zelle',
        'n_in_wait_charge', 'Q_in_charge_wait', 'SOC_in_wait_charge',
        'n_wait', 'Q_wait', 'SOC_wait',
        'n_in_charge', 'Q_in_charge', 'SOC_in_charge',
        'n_in_wait', 'Q_in_wait', 'SOC_in_wait',
        'n_finished_charging', 'Q_finished_charging', 'SOC_finished_charging'
    ]
    sheet.append(headers)

    # Iteriere über alle Zeitschritte und Zellen mit Ladestationen
    for t in model.nb_timestep:
        for c in model.nb_cell:
            # Initialisiere die Summen für die Flotten 10 bis 19
            total_n_in_wait_charge = total_Q_in_charge_wait = total_n_wait = total_Q_wait = 0
            total_n_in_charge = total_Q_in_charge = total_n_in_wait = total_Q_in_wait = 0
            total_n_finished_charging = total_Q_finished_charging = 0

            # Iteriere über die Flotten 10 bis 19 und summiere die Werte
            for fleet_id in range(10, 20):
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    # Erfasse die Werte für diese Flotte
                    Q_in_charge_wait = model.Q_in_charge_wait[t, c, fleet_id].value
                    Q_wait = model.Q_wait[t, c, fleet_id].value
                    Q_in_charge = model.Q_in_charge[t, c, fleet_id].value
                    Q_in_wait = model.Q_in_wait[t, c, fleet_id].value
                    Q_finished_charging = model.Q_finished_charging[t, c, fleet_id].value

                    n_in_wait_charge = model.n_in_wait_charge[t, c, fleet_id].value
                    n_wait = model.n_wait[t, c, fleet_id].value
                    n_in_charge = model.n_in_charge[t, c, fleet_id].value
                    n_in_wait = model.n_in_wait[t, c, fleet_id].value
                    n_finished_charging = model.n_finished_charging[t, c, fleet_id].value

                    # Addiere die Werte zur Gesamtsumme
                    total_n_in_wait_charge += n_in_wait_charge
                    total_Q_in_charge_wait += Q_in_charge_wait
                    total_n_wait += n_wait
                    total_Q_wait += Q_wait
                    total_n_in_charge += n_in_charge
                    total_Q_in_charge += Q_in_charge
                    total_n_in_wait += n_in_wait
                    total_Q_in_wait += Q_in_wait
                    total_n_finished_charging += n_finished_charging
                    total_Q_finished_charging += Q_finished_charging

            # Berechne die SOC-Werte, wenn n-Werte > 0 sind, um Division durch 0 zu vermeiden
            SOC_in_wait_charge = total_Q_in_charge_wait / total_n_in_wait_charge if total_n_in_wait_charge > 0 else 0
            SOC_wait = total_Q_wait / total_n_wait if total_n_wait > 0 else 0
            SOC_in_charge = total_Q_in_charge / total_n_in_charge if total_n_in_charge > 0 else 0
            SOC_in_wait = total_Q_in_wait / total_n_in_wait if total_n_in_wait > 0 else 0
            SOC_finished_charging = total_Q_finished_charging / total_n_finished_charging if total_n_finished_charging > 0 else 0

            # Schreibe die summierten Werte in eine Zeile
            sheet.append([
                t, c,
                total_n_in_wait_charge, total_Q_in_charge_wait, SOC_in_wait_charge,
                total_n_wait, total_Q_wait, SOC_wait,
                total_n_in_charge, total_Q_in_charge, SOC_in_charge,
                total_n_in_wait, total_Q_in_wait, SOC_in_wait,
                total_n_finished_charging, total_Q_finished_charging, SOC_finished_charging
            ])

    # Speichere das Workbook als XLSX-Datei
    workbook.save(file_path)

    print(f"Die kombinierten Lade-Daten der Flotten 10 bis 19 wurden in {file_path} gespeichert.")

def calculate_and_print_energy_split_by_fleet(model: ConcreteModel):
    """
    Berechnet die Gesamtenergie, die von zufälligen und kontrollierten Fahrzeugen geladen wurde, basierend auf E_charge1, E_charge2 und E_charge3.
    Gibt die Ergebnisse in der Konsole aus.

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    """
    total_energy_random = 0
    total_energy_controlled = 0

    controlled_fleet_ids = range(0, 10)
    random_fleet_ids = range(10, 20)

    # Iteriere über die Indizes der Variablen für die Ladeenergie (E_charge1, E_charge2, E_charge3)
    for t in model.nb_timestep:
        for c in model.nb_cell:
            for fleet_id in random_fleet_ids:
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    total_energy_random += (
                        model.E_charge1[t, c, fleet_id].value
                        + model.E_charge2[t, c, fleet_id].value
                        + model.E_charge3[t, c, fleet_id].value
                    )

            for fleet_id in controlled_fleet_ids:
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    total_energy_controlled += (
                        model.E_charge1[t, c, fleet_id].value
                        + model.E_charge2[t, c, fleet_id].value
                        + model.E_charge3[t, c, fleet_id].value
                    )

    # Ausgabe der Ergebnisse in der Konsole
    print(f"Total energy charged (random_fleet): {total_energy_random:.5f} kWh")
    print(f"Total energy charged (controlled_fleet): {total_energy_controlled:.5f} kWh")
    print(f"cross check of total energy charged: {total_energy_random + total_energy_controlled:.5f} kWh")

def calculate_and_print_vehicles_waiting_to_charge_by_fleet(model: ConcreteModel):
    """
    Berechnet die Gesamtanzahl der Fahrzeuge, die bei zufälligen und kontrollierten Flotten auf das Laden warten,
    basierend auf der Variable n_in_wait_charge.
    Gibt die Ergebnisse in der Konsole aus.

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    """
    total_vehicles_waiting_random = 0
    total_vehicles_waiting_controlled = 0

    controlled_fleet_ids = range(0, 10)
    random_fleet_ids = range(10, 20)

    # Iteriere über die Indizes der Variablen für Fahrzeuge, die auf das Laden warten (n_in_wait_charge)
    for t in model.nb_timestep:
        for c in model.nb_cell:
            for fleet_id in random_fleet_ids:
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    total_vehicles_waiting_random += model.n_in_wait_charge[t, c, fleet_id].value

            for fleet_id in controlled_fleet_ids:
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    total_vehicles_waiting_controlled += model.n_in_wait_charge[t, c, fleet_id].value

    # Ausgabe der Ergebnisse in der Konsole
    print(f"Total vehicles to charge (random_fleet): {total_vehicles_waiting_random:.5f}")
    print(f"Total vehicles to charge (controlled_fleet): {total_vehicles_waiting_controlled:.5f}")
    print(f"Cross-check of total vehicles into the CS: {total_vehicles_waiting_random + total_vehicles_waiting_controlled:.5f}")

def calculate_and_print_vehicles_waiting_to_charge_by_fleet_new(model: ConcreteModel):
    """
    Berechnet die Gesamtanzahl der Fahrzeuge, die bei zufälligen und kontrollierten Flotten auf das Laden warten,
    basierend auf der Variable n_in_wait_charge.
    Gibt die Ergebnisse in der Konsole aus, einschließlich der Anzahl pro Flotte (0 bis 19).

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    """
    # Dictionaries, um die Anzahl wartender Fahrzeuge für jede Flotte zu speichern
    fleet_vehicles_waiting = {fleet_id: 0 for fleet_id in range(20)}

    # Iteriere über die Indizes der Variablen für Fahrzeuge, die auf das Laden warten (n_in_wait_charge)
    for t in model.nb_timestep:
        for c in model.nb_cell:
            for fleet_id in range(20):
                if (t, c, fleet_id) in model.charging_cells_key_set:
                    fleet_vehicles_waiting[fleet_id] += model.n_in_wait_charge[t, c, fleet_id].value

    # Berechne die Gesamtzahl der wartenden Fahrzeuge für random und controlled fleets
    total_vehicles_waiting_random = sum(fleet_vehicles_waiting[fleet_id] for fleet_id in range(10, 20))
    total_vehicles_waiting_controlled = sum(fleet_vehicles_waiting[fleet_id] for fleet_id in range(0, 10))

    # Ausgabe der Gesamtanzahlen
    print(f"Total vehicles to charge (random_fleet): {total_vehicles_waiting_random:.5f}")
    print(f"Total vehicles to charge (controlled_fleet): {total_vehicles_waiting_controlled:.5f}")
    print(f"Cross-check of total vehicles into the CS: {total_vehicles_waiting_random + total_vehicles_waiting_controlled:.5f}")

    # Ausgabe der Anzahl wartender Fahrzeuge pro Flotte
    for fleet_id in range(20):
        print(f"Fleet {fleet_id} has {fleet_vehicles_waiting[fleet_id]:.5f} vehicles waiting to charge.")



def write_fleet_energy_and_vehicle_details_to_xlsx_with_soc(model: ConcreteModel, fleet_id: int, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die Energie- und Fahrzeuganzahl einer bestimmten Flotte (fleet_id) über alle Zeitschritte und Zellen ausgibt.
    Zusätzlich wird der State of Charge (SOC) für incoming, in, out und arrived vehicles berechnet und ausgegeben.
    Alle Werte stehen in einer Zeile: n_incoming_vehicles, Q_incoming_vehicles, SOC_incoming_vehicles, n_in, Q_in, SOC_in, etc.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param fleet_id: Die ID der Flotte, für die die Daten ausgegeben werden sollen.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    # Sicherstellen, dass das Verzeichnis 'results' existiert
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    # Erstelle den vollständigen Pfad zur Datei
    file_path = os.path.join(results_dir, filename)

    # Erstelle das Excel-Workbook und aktiviere das Arbeitsblatt
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Fleet_{fleet_id}_Details"

    # Schreibe die Header
    headers = [
        'Zeitschritt', 'Zelle', 'Flotten-id',
        'n_incoming_vehicles', 'Q_incoming_vehicles', 'SOC_incoming_vehicles',
        'n_in', 'Q_in', 'SOC_in',
        'n_out', 'Q_out', 'SOC_out',
        'n_arrived_vehicles', 'Q_arrived_vehicles', 'SOC_arrived_vehicles'
    ]
    sheet.append(headers)

    # Iteriere über alle Zeitschritte und Zellen für die angegebene Flotte
    for t in model.nb_timestep:
        for c in model.nb_cell:
            # Überprüfen, ob die Flotte f in dieser Zelle und diesem Zeitschritt vorhanden ist
            if (t, c, fleet_id) in model.key_set:
                # Erfasse die Werte von Q_in, Q_incoming_vehicles, Q_out, Q_arrived_vehicles
                Q_in = model.Q_in[t, c, fleet_id].value
                Q_incoming_vehicles = model.Q_incoming_vehicles[t, c, fleet_id].value
                Q_out = model.Q_out[t, c, fleet_id].value
                Q_arrived_vehicles = model.Q_arrived_vehicles[t, c, fleet_id].value

                # Erfasse die entsprechenden n-Werte
                n_in = model.n_in[t, c, fleet_id].value
                n_incoming_vehicles = model.n_incoming_vehicles[t, c, fleet_id].value
                n_out = model.n_out[t, c, fleet_id].value
                n_arrived_vehicles = model.n_arrived_vehicles[t, c, fleet_id].value

                # Berechne SOC-Werte, wenn n-Werte > 0 sind, um Division durch 0 zu vermeiden
                SOC_incoming_vehicles = Q_incoming_vehicles / n_incoming_vehicles if n_incoming_vehicles > 0 else 0
                SOC_in = Q_in / n_in if n_in > 0 else 0
                SOC_out = Q_out / n_out if n_out > 0 else 0
                SOC_arrived_vehicles = Q_arrived_vehicles / n_arrived_vehicles if n_arrived_vehicles > 0 else 0

                # Schreibe alle Werte in eine Zeile
                sheet.append([
                    t, c, fleet_id,
                    n_incoming_vehicles, Q_incoming_vehicles, SOC_incoming_vehicles,
                    n_in, Q_in, SOC_in,
                    n_out, Q_out, SOC_out,
                    n_arrived_vehicles, Q_arrived_vehicles, SOC_arrived_vehicles
                ])

    # Speichere das Workbook als XLSX-Datei
    workbook.save(file_path)

    #print(f"Die detaillierten Energie- und Fahrzeug-Daten der Flotte {fleet_id} wurden in {file_path} gespeichert.")


def write_combined_fleet_0_to_9_energy_and_vehicle_details_to_xlsx_with_soc(model: ConcreteModel, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die aggregierte Energie- und Fahrzeuganzahl für die Flotten 0 bis 9 über alle Zeitschritte und Zellen ausgibt.
    Zusätzlich wird der State of Charge (SOC) für incoming, in, out und arrived vehicles berechnet und ausgegeben.
    Alle Werte stehen in einer Zeile: n_incoming_vehicles, Q_incoming_vehicles, SOC_incoming_vehicles, n_in, Q_in, SOC_in, etc.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    file_path = os.path.join(results_dir, filename)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Combined_Fleet_0_to_9_Details"

    headers = [
        'Zeitschritt', 'Zelle',
        'n_incoming_vehicles', 'Q_incoming_vehicles', 'SOC_incoming_vehicles',
        'n_in', 'Q_in', 'SOC_in',
        'n_out', 'Q_out', 'SOC_out',
        'n_arrived_vehicles', 'Q_arrived_vehicles', 'SOC_arrived_vehicles'
    ]
    sheet.append(headers)

    for t in model.nb_timestep:
        for c in model.nb_cell:
            # Initialisiere aggregierte Variablen
            total_n_incoming_vehicles = total_Q_incoming_vehicles = 0
            total_n_in = total_Q_in = 0
            total_n_out = total_Q_out = 0
            total_n_arrived_vehicles = total_Q_arrived_vehicles = 0

            for fleet_id in range(0, 10):
                if (t, c, fleet_id) in model.key_set:
                    total_n_incoming_vehicles += model.n_incoming_vehicles[t, c, fleet_id].value
                    total_Q_incoming_vehicles += model.Q_incoming_vehicles[t, c, fleet_id].value

                    total_n_in += model.n_in[t, c, fleet_id].value
                    total_Q_in += model.Q_in[t, c, fleet_id].value

                    total_n_out += model.n_out[t, c, fleet_id].value
                    total_Q_out += model.Q_out[t, c, fleet_id].value

                    total_n_arrived_vehicles += model.n_arrived_vehicles[t, c, fleet_id].value
                    total_Q_arrived_vehicles += model.Q_arrived_vehicles[t, c, fleet_id].value

            # Berechne SOC-Werte
            SOC_incoming_vehicles = total_Q_incoming_vehicles / total_n_incoming_vehicles if total_n_incoming_vehicles > 0 else 0
            SOC_in = total_Q_in / total_n_in if total_n_in > 0 else 0
            SOC_out = total_Q_out / total_n_out if total_n_out > 0 else 0
            SOC_arrived_vehicles = total_Q_arrived_vehicles / total_n_arrived_vehicles if total_n_arrived_vehicles > 0 else 0

            # Schreibe aggregierte Werte in eine Zeile
            sheet.append([
                t, c,
                total_n_incoming_vehicles, total_Q_incoming_vehicles, SOC_incoming_vehicles,
                total_n_in, total_Q_in, SOC_in,
                total_n_out, total_Q_out, SOC_out,
                total_n_arrived_vehicles, total_Q_arrived_vehicles, SOC_arrived_vehicles
            ])

    workbook.save(file_path)
    print(f"Die aggregierten Energie- und Fahrzeug-Daten für die Flotten 0 bis 9 wurden in {file_path} gespeichert.")

def write_combined_fleet_10_to_19_energy_and_vehicle_details_to_xlsx_with_soc(model: ConcreteModel, filename: str):
    """
    Erzeugt eine XLSX-Datei, die die aggregierte Energie- und Fahrzeuganzahl für die Flotten 10 bis 19 über alle Zeitschritte und Zellen ausgibt.
    Zusätzlich wird der State of Charge (SOC) für incoming, in, out und arrived vehicles berechnet und ausgegeben.
    Alle Werte stehen in einer Zeile: n_incoming_vehicles, Q_incoming_vehicles, SOC_incoming_vehicles, n_in, Q_in, SOC_in, etc.
    Speichert die Datei im Verzeichnis "results".

    :param model: Das ConcreteModel des Ladeoptimierungsmodells.
    :param filename: Der Name der XLSX-Datei, die erzeugt wird (ohne Pfad).
    """
    results_dir = 'results'
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

    file_path = os.path.join(results_dir, filename)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Combined_Fleet_10_to_19_Details"

    headers = [
        'Zeitschritt', 'Zelle',
        'n_incoming_vehicles', 'Q_incoming_vehicles', 'SOC_incoming_vehicles',
        'n_in', 'Q_in', 'SOC_in',
        'n_out', 'Q_out', 'SOC_out',
        'n_arrived_vehicles', 'Q_arrived_vehicles', 'SOC_arrived_vehicles'
    ]
    sheet.append(headers)

    for t in model.nb_timestep:
        for c in model.nb_cell:
            # Initialisiere aggregierte Variablen
            total_n_incoming_vehicles = total_Q_incoming_vehicles = 0
            total_n_in = total_Q_in = 0
            total_n_out = total_Q_out = 0
            total_n_arrived_vehicles = total_Q_arrived_vehicles = 0

            for fleet_id in range(10, 20):
                if (t, c, fleet_id) in model.key_set:
                    total_n_incoming_vehicles += model.n_incoming_vehicles[t, c, fleet_id].value
                    total_Q_incoming_vehicles += model.Q_incoming_vehicles[t, c, fleet_id].value

                    total_n_in += model.n_in[t, c, fleet_id].value
                    total_Q_in += model.Q_in[t, c, fleet_id].value

                    total_n_out += model.n_out[t, c, fleet_id].value
                    total_Q_out += model.Q_out[t, c, fleet_id].value

                    total_n_arrived_vehicles += model.n_arrived_vehicles[t, c, fleet_id].value
                    total_Q_arrived_vehicles += model.Q_arrived_vehicles[t, c, fleet_id].value

            # Berechne SOC-Werte
            SOC_incoming_vehicles = total_Q_incoming_vehicles / total_n_incoming_vehicles if total_n_incoming_vehicles > 0 else 0
            SOC_in = total_Q_in / total_n_in if total_n_in > 0 else 0
            SOC_out = total_Q_out / total_n_out if total_n_out > 0 else 0
            SOC_arrived_vehicles = total_Q_arrived_vehicles / total_n_arrived_vehicles if total_n_arrived_vehicles > 0 else 0

            # Schreibe aggregierte Werte in eine Zeile
            sheet.append([
                t, c,
                total_n_incoming_vehicles, total_Q_incoming_vehicles, SOC_incoming_vehicles,
                total_n_in, total_Q_in, SOC_in,
                total_n_out, total_Q_out, SOC_out,
                total_n_arrived_vehicles, total_Q_arrived_vehicles, SOC_arrived_vehicles
            ])

    workbook.save(file_path)
    print(f"Die aggregierten Energie- und Fahrzeug-Daten für die Flotten 10 bis 19 wurden in {file_path} gespeichert.")
